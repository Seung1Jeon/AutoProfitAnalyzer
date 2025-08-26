import pandas as pd
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

def find_metric_row(sheet, metric_name):
    """주어진 시트의 첫 번째 열에서 특정 지표 이름의 행 번호를 찾습니다."""
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        cell = row[0]
        if cell.value and str(cell.value).strip() == metric_name:
            return cell.row
    return None

def create_profitability_chart(data_sheet, company_name, df_source):
    """수익성 '비율' 지표에 대한 막대 그래프를 생성합니다."""
    first_col_name = df_source.columns[0]
    profitability_ratios = ['매출총이익률', '매출순수익률']
    df_ratios = df_source[df_source[first_col_name].isin(profitability_ratios)]

    if df_ratios.empty:
        return None

    cols_to_include = [col for col in df_ratios.columns if '전년대비' not in str(col)]
    df_for_chart = df_ratios[cols_to_include]

    df_chart = df_for_chart.set_index(df_for_chart.columns[0]).transpose()
    wb = data_sheet.parent
    chart_sheet_name = f"_{company_name}_profit_chart_data"
    if chart_sheet_name in wb.sheetnames:
        wb.remove(wb[chart_sheet_name])
    chart_sheet = wb.create_sheet(chart_sheet_name)
    chart_sheet.sheet_state = 'hidden'

    for r in dataframe_to_rows(df_chart, index=True, header=True):
        chart_sheet.append(r)

    chart = BarChart()
    chart.title = f"{company_name} 수익성 비율(%)"
    chart.x_axis.title = "연도"
    chart.y_axis.title = "비율 (%)"
    chart.x_axis.delete = False

    data = Reference(chart_sheet, min_col=2, min_row=1, max_col=chart_sheet.max_column, max_row=chart_sheet.max_row)
    categories = Reference(chart_sheet, min_col=1, min_row=2, max_col=1, max_row=chart_sheet.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.data_labels = DataLabelList()
    chart.data_labels.showVal = True
    chart.data_labels.dLblPos = 'outEnd'
    chart.height = 10
    chart.width = 20
    return chart

def create_comparison_chart(data_sheet, company_name, metric_name, metric_row, df_source, cpi_ref, categories_ref, bar_color, axis_id):
    """특정 지표와 소비자물가지수를 비교하는 혼합 차트를 생성합니다."""
    metric_series = df_source.iloc[metric_row - 2]
    chart_columns = [col for col in metric_series.index[1:] if '전년대비' not in str(col)]
    chart_values = metric_series[chart_columns].values

    chart_df = pd.DataFrame({
        'Year': chart_columns,
        metric_name: chart_values
    })

    wb = data_sheet.parent
    chart_sheet_name = f"_{company_name}_{metric_name}_comp_data"
    if chart_sheet_name in wb.sheetnames:
        wb.remove(wb[chart_sheet_name])
    chart_sheet = wb.create_sheet(chart_sheet_name)
    chart_sheet.sheet_state = 'hidden'
    for r in dataframe_to_rows(chart_df, index=False, header=True):
        chart_sheet.append(r)

    bar_chart = BarChart()
    bar_chart.title = f"{company_name} {metric_name}과 소비자물가지수"
    bar_chart.x_axis.title = "연도"
    bar_chart.y_axis.title = metric_name
    bar_chart.x_axis.delete = False
    bar_chart.y_axis.position = "l"

    data = Reference(chart_sheet, min_col=2, min_row=1, max_col=2, max_row=chart_sheet.max_row)
    cats = Reference(chart_sheet, min_col=1, min_row=2, max_col=1, max_row=chart_sheet.max_row)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)

    bar_chart.data_labels = DataLabelList()
    bar_chart.data_labels.showVal = True
    bar_chart.series[0].graphicalProperties.solidFill = bar_color

    line_chart = LineChart()
    line_chart.add_data(cpi_ref, titles_from_data=True)
    line_chart.data_labels = DataLabelList()
    line_chart.data_labels.showVal = True
    line_props = line_chart.series[0]
    line_props.graphicalProperties.line.solidFill = "C0504D"
    line_props.marker.symbol = "circle"

    line_chart.y_axis.axId = axis_id
    bar_chart.y_axis.crossAx = axis_id
    bar_chart += line_chart

    bar_chart.sec_axes = [line_chart.y_axis]
    bar_chart.sec_axes[0].position = "r"
    bar_chart.sec_axes[0].title = None
    bar_chart.sec_axes[0].majorGridlines = None
    bar_chart.height = 10
    bar_chart.width = 20
    return bar_chart

def create_single_metric_chart(data_sheet, company_name, metric_name, metric_row, df_source, bar_color):
    """특정 단일 지표에 대한 막대 차트를 생성합니다."""
    metric_series = df_source.iloc[metric_row - 2]
    chart_columns = [col for col in metric_series.index[1:] if '전년대비' not in str(col)]
    chart_values = metric_series[chart_columns].values
    chart_df = pd.DataFrame({
        'Year': chart_columns,
        metric_name: chart_values
    })

    wb = data_sheet.parent
    chart_sheet_name = f"_{company_name}_{metric_name}_data"
    if chart_sheet_name in wb.sheetnames:
        wb.remove(wb[chart_sheet_name])
    chart_sheet = wb.create_sheet(chart_sheet_name)
    chart_sheet.sheet_state = 'hidden'
    for r in dataframe_to_rows(chart_df, index=False, header=True):
        chart_sheet.append(r)

    chart = BarChart()
    chart.title = f"{company_name} {metric_name}"
    chart.x_axis.title = "연도"
    chart.y_axis.title = f"{metric_name} (%)"
    chart.x_axis.delete = False

    data = Reference(chart_sheet, min_col=2, min_row=1, max_col=2, max_row=chart_sheet.max_row)
    cats = Reference(chart_sheet, min_col=1, min_row=2, max_col=1, max_row=chart_sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.data_labels = DataLabelList()
    chart.data_labels.showVal = True
    chart.series[0].graphicalProperties.solidFill = bar_color
    chart.legend = None

    chart.height = 10
    chart.width = 20
    return chart