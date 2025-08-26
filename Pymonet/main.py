import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 기능별로 분리된 모듈 임포트
from py.data_processor import clean_dataframes
from py.data_calculator import add_financial_ratios, calculate_company_average
from py.chart_generator import (
    find_metric_row,
    create_profitability_chart,
    create_comparison_chart,
    create_single_metric_chart
)

def main():
    """메인 실행 함수"""
    # --- 1. 설정 및 데이터 로딩 ---
    filename = 'excels/수익성자동분석.xlsx'
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet1 = wb['Sheet1']
    sheet1.title = "분석결과"

    cpi = pd.read_csv("excels/품목별_소비자물가지수_품목성질별_2020100__20250818162040.csv", encoding='ansi')
    
    DF_LG = pd.read_excel(filename, sheet_name="LG전자")
    DF_Samsung = pd.read_excel(filename, sheet_name="삼성전자")
    DF_SK = pd.read_excel(filename, sheet_name="SK하이닉스")
    Dfs = [DF_LG, DF_Samsung, DF_SK]

    # --- 2. 데이터 정제 및 계산 ---
    Dfs = clean_dataframes(Dfs)
    Dfs[0] = add_financial_ratios(Dfs[0]) # LG전자에 대해서만 비율 추가
    avg_df = calculate_company_average(Dfs)

    # --- 3. 보고서 생성 (테이블, 차트) ---
    
    # 3-1. LG전자 주요 지표 요약 테이블 생성
    df_lg = Dfs[0]
    df_lg_indexed = df_lg.set_index(df_lg.columns[0])
    table_items = ['매출액', '영업이익', '매출총이익률', '매출순수익률', '손익분기점추정']
    table_end_row = 0

    if all(item in df_lg_indexed.index for item in table_items):
        revenue_series = df_lg_indexed.loc['매출액']
        op_profit_series = df_lg_indexed.loc['영업이익']
        gross_margin_series = df_lg_indexed.loc['매출총이익률']
        net_margin_series = df_lg_indexed.loc['매출순수익률']
        bep_series = df_lg_indexed.loc['손익분기점추정']
        year_cols = [col for col in df_lg.columns[1:] if '전년대비' not in str(col)]
        
        summary_data = []
        for year in year_cols:
            revenue_value = revenue_series.get(year, 0)
            bep_value = bep_series.get(year, 0)
            status = '해당없음'
            if not (pd.isna(bep_value) or bep_value == 0):
                status = '달성' if revenue_value >= bep_value else '미달'
            
            summary_data.append({
                '연도': year, '매출액': revenue_value, '영업이익': op_profit_series.get(year, 0),
                '매출총이익률(%)': gross_margin_series.get(year, 0), '매출순수익률(%)': net_margin_series.get(year, 0),
                '손익분기점 추정': bep_value, '달성여부': status
            })
        summary_df = pd.DataFrame(summary_data)
        
        sheet1.cell(row=1, column=1, value="LG전자 주요 지표 요약").font = Font(bold=True, size=14)
        rows = dataframe_to_rows(summary_df, index=False, header=True)
        for r_idx, row in enumerate(rows, 3):
            for c_idx, value in enumerate(row, 1):
                cell = sheet1.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        table_end_row = 3 + len(summary_df)

    # 3-2. 정제된 데이터를 새 시트에 다시 쓰기
    sheet_names = ["LG전자", "삼성전자", "SK하이닉스"]
    company_sheets = {}
    for df, sheet_name in zip(Dfs, sheet_names):
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        new_sheet = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)
        company_sheets[sheet_name] = new_sheet
    
    sheet_LG = company_sheets["LG전자"]

    # 3-3. CPI 데이터 및 차트 위치 설정
    cpi_start_row = table_end_row + 2 if table_end_row > 0 else 1
    sheet1.cell(row=cpi_start_row, column=1, value='연도')
    sheet1.cell(row=cpi_start_row, column=2, value='소비자물가지수')
    cpi_data_rows = [['2020', np.mean(cpi['2020'])], ['2021', np.mean(cpi['2021'])], ['2022', np.mean(cpi['2022'])], ['2023', np.mean(cpi['2023'])], ['2024', np.mean(cpi['2024'])]]
    for r_offset, row_data in enumerate(cpi_data_rows):
        sheet1.cell(row=cpi_start_row + 1 + r_offset, column=1, value=row_data[0])
        sheet1.cell(row=cpi_start_row + 1 + r_offset, column=2, value=row_data[1])
    cpi_end_row = cpi_start_row + len(cpi_data_rows)
    charts_start_row = cpi_end_row + 2

    # 3-4. 차트 생성 및 추가
    data_cpi_ref = Reference(sheet1, min_col=2, min_row=cpi_start_row, max_col=2, max_row=cpi_end_row)
    categories_ref = Reference(sheet1, min_col=1, min_row=cpi_start_row + 1, max_col=1, max_row=cpi_end_row)

    # CPI 단독 차트
    chart_cpi_only = LineChart()
    chart_cpi_only.title = "연도별 소비자물가지수"
    chart_cpi_only.add_data(data_cpi_ref, titles_from_data=True)
    chart_cpi_only.set_categories(categories_ref)
    chart_cpi_only.height, chart_cpi_only.width = 10, 20
    sheet1.add_chart(chart_cpi_only, f"A{charts_start_row}")

    # LG전자 수익성 비율 차트
    profitability_chart_lg = create_profitability_chart(sheet_LG, "LG전자", Dfs[0])
    if profitability_chart_lg:
        sheet1.add_chart(profitability_chart_lg, f"M{charts_start_row}")

    # LG전자 매출액/영업이익 vs CPI 차트
    revenue_row_lg = find_metric_row(sheet_LG, "매출액")
    if revenue_row_lg:
        revenue_chart_lg = create_comparison_chart(sheet_LG, "LG전자", "매출액", revenue_row_lg, Dfs[0], data_cpi_ref, categories_ref, "4F81BD", 200)
        sheet1.add_chart(revenue_chart_lg, f"A{charts_start_row + 20}")
    
    op_profit_row_lg = find_metric_row(sheet_LG, "영업이익")
    if op_profit_row_lg:
        op_profit_chart_lg = create_comparison_chart(sheet_LG, "LG전자", "영업이익", op_profit_row_lg, Dfs[0], data_cpi_ref, categories_ref, "9BBB59", 201)
        sheet1.add_chart(op_profit_chart_lg, f"M{charts_start_row + 20}")

    # LG전자 vs 3사 평균 비교 차트
    if avg_df is not None:
        df_lg_indexed = Dfs[0].set_index(Dfs[0].columns[0])
        avg_df_indexed = avg_df.set_index(avg_df.columns[0])
        
        if '매출액' in df_lg_indexed.index and '3사 평균 매출액' in avg_df_indexed.index:
            lg_revenue = df_lg_indexed.loc[['매출액']]; lg_revenue.index = ['LG전자 매출액']
            avg_revenue = avg_df_indexed.loc[['3사 평균 매출액']]
            revenue_comp_df = pd.concat([avg_revenue, lg_revenue]).reindex(columns=avg_revenue.columns)
            revenue_comp_df_t = revenue_comp_df[[c for c in revenue_comp_df.columns if '전년대비' not in str(c)]].transpose()
            comp_rev_sheet = wb.create_sheet("매출액_비교_데이터")
            for r in dataframe_to_rows(revenue_comp_df_t, index=True, header=True): comp_rev_sheet.append(r)
            chart_comp_rev = BarChart()
            chart_comp_rev.title = "매출액 비교 (LG전자 vs 3사 평균)"
            data = Reference(comp_rev_sheet, min_col=2, min_row=1, max_col=comp_rev_sheet.max_column, max_row=comp_rev_sheet.max_row)
            cats = Reference(comp_rev_sheet, min_col=1, min_row=2, max_col=1, max_row=comp_rev_sheet.max_row)
            chart_comp_rev.add_data(data, titles_from_data=True); chart_comp_rev.set_categories(cats)
            chart_comp_rev.height, chart_comp_rev.width = 10, 20
            sheet1.add_chart(chart_comp_rev, f"A{charts_start_row + 40}")

        if '영업이익' in df_lg_indexed.index and '3사 평균 영업이익' in avg_df_indexed.index:
            lg_op_profit = df_lg_indexed.loc[['영업이익']]; lg_op_profit.index = ['LG전자 영업이익']
            avg_op_profit = avg_df_indexed.loc[['3사 평균 영업이익']]
            op_profit_comp_df = pd.concat([avg_op_profit, lg_op_profit]).reindex(columns=avg_op_profit.columns)
            op_profit_comp_df_t = op_profit_comp_df[[c for c in op_profit_comp_df.columns if '전년대비' not in str(c)]].transpose()
            comp_op_sheet = wb.create_sheet("영업이익_비교_데이터")
            for r in dataframe_to_rows(op_profit_comp_df_t, index=True, header=True): comp_op_sheet.append(r)
            chart_comp_op = BarChart()
            chart_comp_op.title = "영업이익 비교 (LG전자 vs 3사 평균)"
            data = Reference(comp_op_sheet, min_col=2, min_row=1, max_col=comp_op_sheet.max_column, max_row=comp_op_sheet.max_row)
            cats = Reference(comp_op_sheet, min_col=1, min_row=2, max_col=1, max_row=comp_op_sheet.max_row)
            chart_comp_op.add_data(data, titles_from_data=True); chart_comp_op.set_categories(cats)
            chart_comp_op.height, chart_comp_op.width = 10, 20
            sheet1.add_chart(chart_comp_op, f"M{charts_start_row + 40}")

    # LG전자 개별 비율 차트
    net_profit_margin_row_lg = find_metric_row(sheet_LG, "매출순수익률")
    if net_profit_margin_row_lg:
        net_profit_margin_chart_lg = create_single_metric_chart(sheet_LG, "LG전자", "매출순수익률", net_profit_margin_row_lg, Dfs[0], "FFC000")
        sheet1.add_chart(net_profit_margin_chart_lg, f"A{charts_start_row + 60}")

    gross_profit_margin_row_lg = find_metric_row(sheet_LG, "매출총이익률")
    if gross_profit_margin_row_lg:
        gross_profit_margin_chart_lg = create_single_metric_chart(sheet_LG, "LG전자", "매출총이익률", gross_profit_margin_row_lg, Dfs[0], "7030A0")
        sheet1.add_chart(gross_profit_margin_chart_lg, f"M{charts_start_row + 60}")

    # --- 4. 파일 저장 ---
    wb.save("excels/수익성_분석결과.xlsx")

if __name__ == "__main__":
    main()
